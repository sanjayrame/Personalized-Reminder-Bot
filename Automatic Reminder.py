import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value
unbookedMembers = {}
for r in range(2, sheet.max_row + 1):
    booking = sheet.cell(row=r, column=lastCol).value
    if booking != 'booked':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unbookedMembers[name] = email

smtpObj = smtplib.SMTP('smtp.sanjay.com', 684)

smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('sanjayrame@gmail.com', sys.argv[1])




for name, email in unbookedMembers.items():
    body = "Subject: Book chess classes for %s .\nHey %s!\n This is a friendly reminder that you forgot to schedule your classes for %s. Feel free to schedule them at your convenience. Thank you!'" %(latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('sanjayrame@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))


smtpObj.quit()